"""
Export endpoints — genera archivos descargables.
"""
import io
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_db
from models import HistorialEntry, Causacion
import json

router = APIRouter(prefix="/export", tags=["export"])


def _currency(n: float) -> str:
    return f"$ {n:,.0f}".replace(",", ".")


def _style_header(ws, row: int, cols: int):
    fill   = PatternFill("solid", fgColor="0D1A30")
    font   = Font(bold=True, color="00D97E", name="Calibri", size=10)
    border = Border(bottom=Side(style="thin", color="1E3358"))
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill, cell.font, cell.border = fill, font, border
        cell.alignment = Alignment(horizontal="center", vertical="center")


@router.get("/historial")
def export_historial(
    status:   str | None = Query(None),
    platform: str | None = Query(None),
    search:   str | None = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(HistorialEntry).order_by(HistorialEntry.id.desc())
    if status:   q = q.filter(HistorialEntry.status   == status)
    if platform: q = q.filter(HistorialEntry.platform == platform)
    if search:   q = q.filter(HistorialEntry.concepto.ilike(f"%{search}%"))
    rows = q.limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Contable"

    # Título
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "ContaFlow — Historial de Asientos Contables"
    title_cell.font  = Font(bold=True, size=14, color="00D97E", name="Calibri")
    title_cell.fill  = PatternFill("solid", fgColor="080F1E")
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["ID", "Fecha", "Hora", "Concepto", "Tipo", "Plataforma", "Cuentas", "Débito (COP)", "Crédito (COP)", "Estado"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header(ws, 2, len(headers))
    ws.row_dimensions[2].height = 20

    status_colors = {"ok": "00D97E", "pending": "F0A500", "error": "FF4757"}
    for r, entry in enumerate(rows, 3):
        data = [entry.id, entry.fecha, entry.hora, entry.concepto, entry.tipo,
                entry.platform.upper(), entry.cuentas,
                entry.debito, entry.credito, entry.status.upper()]
        fill_color = "112040" if r % 2 == 0 else "0D1A30"
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(color="E8EEF8", name="Calibri", size=9)
            cell.alignment = Alignment(vertical="center")
            if col in (8, 9):
                cell.number_format = '#,##0'
                cell.font = Font(color="00D97E", name="Calibri", size=9)
            if col == 10:
                color = status_colors.get(entry.status, "7A95B8")
                cell.font = Font(color=color, bold=True, name="Calibri", size=9)

    col_widths = [6, 12, 7, 35, 15, 12, 9, 18, 18, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Totales
    total_row = len(rows) + 3
    ws.cell(total_row, 4, "TOTAL").font = Font(bold=True, color="E8EEF8", name="Calibri")
    ws.cell(total_row, 8, sum(r.debito  for r in rows)).number_format = '#,##0'
    ws.cell(total_row, 9, sum(r.credito for r in rows)).number_format = '#,##0'
    for col in (4, 8, 9):
        ws.cell(total_row, col).fill = PatternFill("solid", fgColor="1A3060")
        ws.cell(total_row, col).font = Font(bold=True, color="00D97E", name="Calibri")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"contaflow_historial_{__import__('datetime').date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/causaciones")
def export_causaciones_list(
    status:   str | None = Query(None),
    platform: str | None = Query(None),
    search:   str | None = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(Causacion).order_by(Causacion.id.desc())
    if status:   q = q.filter(Causacion.status   == status)
    if platform: q = q.filter(Causacion.plataforma == platform)
    if search:   q = q.filter(Causacion.concepto.ilike(f"%{search}%"))
    rows = q.limit(5000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Causaciones"

    ws.merge_cells("A1:I1")
    ws["A1"].value = "ContaFlow — Causaciones Contables"
    ws["A1"].font  = Font(bold=True, size=14, color="00D97E", name="Calibri")
    ws["A1"].fill  = PatternFill("solid", fgColor="080F1E")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["ID", "Fecha", "Concepto", "Plataforma", "Líneas", "Débito (COP)", "Crédito (COP)", "Estado", "Usuario"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    _style_header(ws, 2, len(headers))
    ws.row_dimensions[2].height = 20

    status_colors = {"ok": "00D97E", "pending": "F0A500", "error": "FF4757"}
    for r, caus in enumerate(rows, 3):
        lines = json.loads(caus.lineas or "[]")
        data = [caus.id, caus.fecha, caus.concepto, caus.plataforma.upper(),
                len(lines), caus.total_debito, caus.total_credito, caus.status.upper(), caus.usuario]
        fill_color = "112040" if r % 2 == 0 else "0D1A30"
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.font = Font(color="E8EEF8", name="Calibri", size=9)
            cell.alignment = Alignment(vertical="center")
            if col in (6, 7):
                cell.number_format = '#,##0'
                cell.font = Font(color="00D97E", name="Calibri", size=9)
            if col == 8:
                color = status_colors.get(caus.status, "7A95B8")
                cell.font = Font(color=color, bold=True, name="Calibri", size=9)

    col_widths = [6, 12, 38, 12, 8, 18, 18, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total_row = len(rows) + 3
    ws.cell(total_row, 3, "TOTAL").font = Font(bold=True, color="E8EEF8", name="Calibri")
    ws.cell(total_row, 6, sum(c.total_debito  for c in rows)).number_format = '#,##0'
    ws.cell(total_row, 7, sum(c.total_credito for c in rows)).number_format = '#,##0'
    for col in (3, 6, 7):
        ws.cell(total_row, col).fill = PatternFill("solid", fgColor="1A3060")
        ws.cell(total_row, col).font = Font(bold=True, color="00D97E", name="Calibri")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"contaflow_causaciones_{__import__('datetime').date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/causacion/{causacion_id}")
def export_causacion(causacion_id: int, db: Session = Depends(get_db)):
    caus = db.query(Causacion).filter_by(id=causacion_id).first()
    if not caus:
        from fastapi import HTTPException
        raise HTTPException(404, "No encontrado")

    lineas = json.loads(caus.lineas or "[]")
    wb = Workbook()
    ws = wb.active
    ws.title = "Asiento Contable"

    ws.merge_cells("A1:F1")
    ws["A1"].value = f"Asiento #{caus.id} — {caus.concepto}"
    ws["A1"].font  = Font(bold=True, size=13, color="00D97E", name="Calibri")
    ws["A1"].fill  = PatternFill("solid", fgColor="080F1E")
    ws["A1"].alignment = Alignment(horizontal="center")

    meta = [("Fecha", caus.fecha), ("Plataforma", caus.plataforma.upper()),
            ("Estado", caus.status.upper()), ("Usuario", caus.usuario)]
    for i, (k, v) in enumerate(meta, 2):
        ws.cell(i, 1, k).font = Font(bold=True, color="7A95B8", name="Calibri", size=9)
        ws.cell(i, 2, v).font = Font(color="E8EEF8", name="Calibri", size=9)

    headers = ["Cuenta PUC", "Descripción", "Tercero / NIT", "Débito (COP)", "Crédito (COP)"]
    start = 7
    for col, h in enumerate(headers, 1):
        ws.cell(start, col, h)
    _style_header(ws, start, len(headers))

    for r, l in enumerate(lineas, start + 1):
        ws.cell(r, 1, l.get("cuenta","")).font = Font(color="4A9EFF", name="Calibri", size=9)
        ws.cell(r, 2, l.get("descripcion","")).font = Font(color="E8EEF8", name="Calibri", size=9)
        ws.cell(r, 3, l.get("tercero","")).font = Font(color="7A95B8", name="Calibri", size=9)
        ws.cell(r, 4, l.get("debito",0)).number_format = '#,##0'
        ws.cell(r, 5, l.get("credito",0)).number_format = '#,##0'
        fill = PatternFill("solid", fgColor="112040" if r % 2 == 0 else "0D1A30")
        for c in range(1, 6): ws.cell(r, c).fill = fill

    for col, w in enumerate([14, 30, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="asiento_{causacion_id}.xlsx"'},
    )
